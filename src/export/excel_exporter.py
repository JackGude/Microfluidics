from __future__ import annotations

import os
from typing import Dict, List, Sequence

import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm  # type: ignore
except Exception:  # pragma: no cover
    tqdm = None


def _is_fixed_batch_array(values: object, n_rows: int) -> bool:
    return (
        isinstance(values, np.ndarray)
        and values.ndim in (1, 2)
        and values.shape[0] == n_rows
        and values.dtype != object
    )


def _sanitize_sheet_name(name: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if c in invalid else c for c in name).strip()
    if not cleaned:
        cleaned = "sheet"
    return cleaned[:31]


def _unique_sheet_name(base: str, used: set[str]) -> str:
    if base not in used:
        used.add(base)
        return base
    i = 2
    while True:
        candidate = f"{base[:28]}_{i}" if len(base) > 28 else f"{base}_{i}"
        candidate = candidate[:31]
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def _as_2d(values: np.ndarray) -> np.ndarray:
    if values.ndim == 1:
        return values.reshape(values.shape[0], 1)
    return values


def _pad_ragged_vectors(values: Sequence[object], n_rows: int) -> np.ndarray:
    max_len = 0
    parsed: List[np.ndarray] = []
    for v in values:
        if isinstance(v, np.ndarray):
            arr = v
        else:
            arr = np.asarray(v)
        if arr.ndim != 1:
            raise ValueError("Per-row feature vectors must be 1D")
        parsed.append(arr)
        if arr.shape[0] > max_len:
            max_len = arr.shape[0]

    out = np.full((n_rows, max_len), np.nan, dtype=float)
    for i, arr in enumerate(parsed):
        out[i, : arr.shape[0]] = arr.astype(float)
    return out


def export_features_to_excel(
    *,
    labels: np.ndarray,
    feature_dicts: Sequence[Dict[str, object]],
    output_path: str,
) -> None:
    if not isinstance(labels, np.ndarray):
        raise TypeError("labels must be a numpy array")
    if labels.ndim != 1:
        raise ValueError("labels must be a 1D array")

    n_rows = labels.shape[0]

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    used_names: set[str] = set()

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)

    feature_items: List[tuple[str, object]] = []
    for feature_dict in feature_dicts:
        for feature_name, values in feature_dict.items():
            feature_items.append((feature_name, values))

    feature_iter = feature_items
    if tqdm is not None:
        feature_iter = tqdm(feature_items, desc="Features", unit="sheet")

    for feature_name, values in feature_iter:
        if _is_fixed_batch_array(values, n_rows):
            data = _as_2d(values)
        elif isinstance(values, (list, tuple)) and len(values) == n_rows:
            data = _pad_ragged_vectors(values, n_rows)
        else:
            raise ValueError(
                f"Unsupported feature output for '{feature_name}'. "
                "Use a fixed numpy array with shape (n_rows,) or (n_rows, n_features), "
                "or a per-row list/tuple of 1D vectors."
            )

        base = _sanitize_sheet_name(str(feature_name))
        sheet_name = _unique_sheet_name(base, used_names)
        ws = wb.create_sheet(title=sheet_name)

        n_cols = int(data.shape[1])

        ws.cell(row=1, column=1, value="row_id")
        ws.cell(row=1, column=2, value="label")

        first_feature_col = 3
        last_feature_col = first_feature_col + n_cols - 1
        if n_cols == 1:
            ws.cell(row=1, column=first_feature_col, value=str(feature_name))
        else:
            ws.merge_cells(
                start_row=1,
                start_column=first_feature_col,
                end_row=1,
                end_column=last_feature_col,
            )
            ws.cell(row=1, column=first_feature_col, value=str(feature_name))

        row_iter = range(n_rows)
        if tqdm is not None:
            row_iter = tqdm(row_iter, desc=f"Writing {sheet_name}", unit="row", leave=False)

        for i in row_iter:
            excel_row = i + 2
            ws.cell(row=excel_row, column=1, value=i)
            ws.cell(row=excel_row, column=2, value=int(labels[i]))
            row_vals = data[i]
            for j in range(n_cols):
                v = row_vals[j]
                ws.cell(row=excel_row, column=first_feature_col + j, value=None if np.isnan(v) else float(v))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(last_feature_col)}{n_rows + 1}"

    wb.save(output_path)
